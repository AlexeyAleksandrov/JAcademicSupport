#!/usr/bin/env python3
"""
import_expert_opinions.py
Reads ТОП ИТ 2026 Excel files and imports technologies from "Примеры технологий"
into the expert_opinion table in AcademicSupport DB.

Usage:
  python import_expert_opinions.py            # run import
  python import_expert_opinions.py --dry-run  # preview only, no DB writes
"""

import sys
import re
import io
import openpyxl
import psycopg2
from pathlib import Path

# Force UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Config ────────────────────────────────────────────────────────────────────

DB_PARAMS = dict(host="localhost", port=5432, dbname="AcademicSupport",
                 user="postgres", password="1111")

BASE_DIR = Path(r"O:\МИРЭА\Учёба\Диссертация\ТОП ИТ 2026")

# file → profession_code  (None = universal / cross-cutting)
FILES = {
    "Инженер-программист.xlsx":          "engineer-programmer",
    "Продакт-менеджер.xlsx":             "product-manager",
    "Руководитель группой разработки.xlsx": "dev-team-lead",
    "Перекрестные компетенции.xlsx":     None,
}

LEVEL_IMPORTANCE = {
    "базовый":     0.65,
    "средний":     0.80,
    "продвинутый": 0.90,
}

DEFAULT_IMPORTANCE = 0.75

DRY_RUN = "--dry-run" in sys.argv


# ── Text helpers ──────────────────────────────────────────────────────────────

def clean_tech_name(s: str) -> str:
    s = s.strip()
    s = re.sub(r'^[\s\-–•*·◦▸▪\d\.]+', '', s).strip()   # leading bullets/numbers
    s = re.sub(r'[;,\.]+$', '', s).strip()                # trailing punctuation
    return s


def extract_technologies(cell_text: str) -> list:
    """
    Parse multi-line, comma-separated cell content into individual technology names.
    """
    if not cell_text:
        return []

    # Category prefixes to strip (short non-tech labels before colon)
    NON_TECH_PREFIXES = re.compile(
        r'^(инструмент[ыи]?|библиотек[аи]?|фреймворк[ыи]?|среды?|платформ[аы]?|'
        r'системн\w+|языки?|методы?|cms|api|ос|по|среда разработки|'
        r'сервисы?|инструментарий|визуализация|профилировщики?|'
        r'моделирование|форматы?|стандарты?|синтаксис|технологии|концепции?)\s*$',
        re.IGNORECASE
    )

    # Compound names that should NOT be split by slash
    KEEP_SLASH = re.compile(r'^(c/c\+\+|html/css|tcp/ip|http/https|ci/cd|read/write)$', re.IGNORECASE)

    text = str(cell_text)
    techs = []

    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue

        # Strip category prefix ("CMS: Tilda" → "Tilda")
        m = re.match(r'^([^:]{2,25}):\s+(.+)$', line)
        if m:
            prefix = m.group(1).strip()
            # Strip if prefix matches non-tech pattern OR is a very short label (≤5 chars)
            if NON_TECH_PREFIXES.match(prefix) or len(prefix) <= 5:
                line = m.group(2).strip()

        # Split by comma
        parts = split_by_comma(line)

        for part in parts:
            # Split by " и " / " and "
            sub_parts = re.split(r'\s+(?:и|and)\s+', part, flags=re.IGNORECASE)
            for sub in sub_parts:
                name = clean_tech_name(sub)
                if not name or len(name) < 2:
                    continue

                # Split by "/" unless it's a known compound
                if '/' in name and not KEEP_SLASH.match(name):
                    slash_parts = [s.strip() for s in name.split('/')]
                else:
                    slash_parts = [name]

                for sp in slash_parts:
                    sp = clean_tech_name(sp)
                    if not sp or len(sp) < 2:
                        continue

                    name_base = re.sub(r'\s+v?\d[\d\.x\-\+]*$', '', sp).strip()
                    name_no_paren = re.sub(r'\s*\([^)]*\)\s*$', '', name_base).strip()

                    candidates = []
                    if name_no_paren and name_no_paren != name_base:
                        candidates.append(name_no_paren)
                    candidates.append(name_base)
                    if name_base != sp:
                        candidates.append(sp)

                    techs.append(candidates)

    # Deduplicate by first candidate
    seen = set()
    result = []
    for c in techs:
        key = c[0].lower()
        if key not in seen:
            seen.add(key)
            result.append(c)
    return result


def split_by_comma(text: str) -> list:
    """Split by comma but not inside parentheses."""
    parts = []
    depth = 0
    cur = []
    for ch in text:
        if ch == '(':
            depth += 1
            cur.append(ch)
        elif ch == ')':
            depth -= 1
            cur.append(ch)
        elif ch == ',' and depth == 0:
            parts.append(''.join(cur).strip())
            cur = []
        else:
            cur.append(ch)
    if cur:
        parts.append(''.join(cur).strip())
    return [p for p in parts if p]


# ── Excel helpers ─────────────────────────────────────────────────────────────

def find_header_info(ws):
    """Return (header_row_idx, col_map) where col_map maps field → column index."""
    for row_idx in range(1, 8):
        row = ws[row_idx]
        col_map = {}
        for cell in row:
            val = str(cell.value or "").strip()
            if "Примеры технологий" in val:
                col_map["tech"] = cell.column
            if "Источник" in val:
                col_map["source"] = cell.column
            if re.search(r'Уровень', val):
                col_map["level"] = cell.column
            if re.search(r'Компетен', val, re.I):
                col_map["competency"] = cell.column
        if "tech" in col_map:
            return row_idx, col_map
    return 1, {}


def merged_value(ws, row_idx: int, col_idx, cache: dict):
    """Get cell value, using last non-None from above (handles merged rows)."""
    if col_idx is None:
        return cache.get(col_idx)
    cell = ws.cell(row=row_idx, column=col_idx)
    val = cell.value
    if val is not None:
        cache[col_idx] = val
    return cache.get(col_idx)


# ── DB helpers ────────────────────────────────────────────────────────────────

def get_or_create_expert(cur, name: str, trust: float = 0.85) -> int:
    name = re.sub(r'\s+', ' ', str(name).strip())[:255]
    cur.execute("SELECT id FROM expert WHERE name = %s", (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    if DRY_RUN:
        print(f"  [DRY] New expert: {name!r}")
        return -1
    cur.execute("INSERT INTO expert (name, trust) VALUES (%s, %s) RETURNING id",
                (name, trust))
    new_id = cur.fetchone()[0]
    print(f"  + Expert #{new_id}: {name!r}")
    return new_id


def lookup_canonical(cur, candidates: list):
    """Try each candidate name and return (canonical_id, domain, tech_family) or None."""
    for name in candidates:
        if not name or len(name) < 2:
            continue
        cur.execute(
            "SELECT id, domain, tech_family FROM skill_canonical WHERE LOWER(name) = LOWER(%s) LIMIT 1",
            (name,)
        )
        row = cur.fetchone()
        if row:
            return row
    # Second pass: partial match on normalized_name
    for name in candidates:
        norm = re.sub(r'[^a-zа-яё0-9]', '', name.lower())
        if len(norm) < 3:
            continue
        cur.execute(
            "SELECT id, domain, tech_family FROM skill_canonical "
            "WHERE LOWER(REGEXP_REPLACE(normalized_name, '[^a-zа-яёa-z0-9]', '', 'gi')) = %s LIMIT 1",
            (norm,)
        )
        row = cur.fetchone()
        if row:
            return row
    return None


def already_exists(cur, expert_id, canonical_id, profession_code) -> bool:
    cur.execute(
        "SELECT 1 FROM expert_opinion "
        "WHERE expert_id=%s AND canonical_id=%s AND profession_code IS NOT DISTINCT FROM %s",
        (expert_id, canonical_id, profession_code)
    )
    return cur.fetchone() is not None


# ── Core processing ───────────────────────────────────────────────────────────

def process_file(conn, filename: str, profession_code):
    path = BASE_DIR / filename
    print(f"\n{'─'*60}")
    print(f"File:    {filename}")
    print(f"Profcode:{profession_code}")
    print(f"{'─'*60}")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    header_row, col_map = find_header_info(ws)
    print(f"Header row: {header_row}")
    print(f"Columns:    {col_map}")

    if "tech" not in col_map:
        print("  WARNING: 'Примеры технологий' column not found — skipping")
        return 0

    cache = {}
    inserted = skipped_dup = skipped_no_canonical = 0
    not_found_techs = set()

    with conn.cursor() as cur:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            level_val  = merged_value(ws, row_idx, col_map.get("level"),  cache)
            source_val = merged_value(ws, row_idx, col_map.get("source"), cache)
            tech_cell  = ws.cell(row=row_idx, column=col_map["tech"]).value

            if not tech_cell:
                continue

            # Skill importance from level
            level_str  = str(level_val or "").strip().lower()
            importance = LEVEL_IMPORTANCE.get(level_str, DEFAULT_IMPORTANCE)

            # Expert from source — take first non-empty line only
            raw_source = str(source_val or "").strip()
            first_line = next((ln.strip() for ln in raw_source.split('\n') if ln.strip()), "")
            source_name = re.sub(r'\s+', ' ', first_line) or "Рекомендации работодателей"
            expert_id = get_or_create_expert(cur, source_name)

            # Parse and insert each technology
            tech_list = extract_technologies(str(tech_cell))

            for candidates in tech_list:
                canonical = lookup_canonical(cur, candidates)
                if canonical is None:
                    skipped_no_canonical += 1
                    not_found_techs.add(candidates[0])
                    continue

                canonical_id, domain, tech_family = canonical

                if not DRY_RUN and already_exists(cur, expert_id, canonical_id, profession_code):
                    skipped_dup += 1
                    continue

                if DRY_RUN:
                    print(f"    [DRY] {candidates[0]!r} → canonical #{canonical_id}, "
                          f"domain={domain}, importance={importance:.2f}")
                    inserted += 1
                    continue

                cur.execute(
                    """INSERT INTO expert_opinion
                         (expert_id, skill_importance, canonical_id, direction,
                          profession_code, domain, tech_family)
                       VALUES (%s, %s, %s, 'POSITIVE', %s, %s, %s)""",
                    (expert_id, importance, canonical_id, profession_code, domain, tech_family)
                )
                inserted += 1

        if not DRY_RUN:
            conn.commit()

    print(f"\nResults → inserted: {inserted}, duplicates skipped: {skipped_dup}, "
          f"no canonical: {skipped_no_canonical}")
    if not_found_techs:
        print(f"Not found in skill_canonical ({len(not_found_techs)}):")
        for t in sorted(not_found_techs)[:30]:
            print(f"  - {t!r}")
    return inserted


def main():
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    if DRY_RUN:
        print("=== DRY RUN MODE ===")

    conn = psycopg2.connect(**DB_PARAMS)
    total = 0

    for filename, profession_code in FILES.items():
        try:
            total += process_file(conn, filename, profession_code)
        except FileNotFoundError:
            print(f"\nERROR: File not found: {BASE_DIR / filename}")
        except Exception as e:
            import traceback
            print(f"\nERROR in {filename}: {e}")
            traceback.print_exc()

    conn.close()
    print(f"\n{'═'*60}")
    print(f"TOTAL inserted: {total}")


if __name__ == "__main__":
    main()
